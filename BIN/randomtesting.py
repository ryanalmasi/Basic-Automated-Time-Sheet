# imports 
from errno import EREMOTE
from os import times
from pathlib import Path
from re import T
from types import NoneType
import webbrowser
from numpy import true_divide
import pandas as pd
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import glob


# initializing every month regardless of what haley enters
input_file = "TS-2022_AbhroChowdhury.xlsx"
timesheet_wb = load_workbook("TS-2022_AbhroChowdhury.xlsx")   # loading input excel book
output_file = 'OT_template.xlsx'    # saving the output file name for later
output_wb = load_workbook(output_file)   # loading in the output excel book

#Initializing every sheet
#January:
january_input_sheet = timesheet_wb["Jan"]
january_output_sheet = output_wb["JAN 22"]

#February:
february_input_sheet = timesheet_wb["Feb"]
february_output_sheet = output_wb["FEB 22"]

#March:
march_input_sheet = timesheet_wb["Mar"]
march_output_sheet = output_wb["MAR 22"]

#April
april_input_sheet = timesheet_wb["Apr"]
april_output_sheet = output_wb["APR 22"]

#May 
may_input_sheet = timesheet_wb["May"]
may_output_sheet = output_wb["MAY 22"]

#June 
june_input_sheet = timesheet_wb["Jun"]
june_output_sheet = output_wb["JUN 22"]

#July
july_input_sheet = timesheet_wb["Jul"]
july_output_sheet = output_file["JUL 22"]

#August 
july_input_sheet = timesheet_wb["Aug"]
july_output_sheet = output_wb["AUG 22"]

#September
sept_input_sheet = timesheet_wb["Sep"]
sept_output_sheet = output_wb["SEP 22"]

#October
oct_input_sheet = timesheet_wb["Oct"]
oct_output_sheet = output_wb["OCT 22"]

#November
nov_input_sheet = timesheet_wb["Nov"]
nov_output_sheet = output_wb["NOV 22"]

#December:
december_input_sheet = timesheet_wb["Dec"]
december_output_sheet = output_wb["DEC 22"]

inputlist = []
inputlist.append(december_input_sheet)
inputlist.append(july_input_sheet)

outputlist = []
outputlist.append(december_output_sheet)
outputlist.append(july_output_sheet)

#function for OTperday
def OTperday(specialcolumn):
    mylist = []    # creating an empty list to later append values to 
    for i in range(11, 30):    # range of rows that I want to iterate over (usually 11 to 70)
        checktype = type(month.cell(row=i, column=specialcolumn).value)  # this will pick out the specific cell, and check the type of value in it
        if checktype == int:   # I want to only add integers to my list
            data = month.cell(row=i, column=specialcolumn).value     # gives me value for the column we want, and will iterate over each row in that column
            mylist.append(data)    # adding values into the list, this will give us the final overtime hours in that particular day
    thesum = sum(mylist)    # taking total of the list to calculate total OT for that particular day
    return thesum

#to scan through each months sheet and to generate an OTlist for that month
for themonth in inputlist:
    if themonth == december_input_sheet:
        month = december_input_sheet
        OTlist = []
        for d in range(1, 31):
            OTdata = OTperday(4 + d*2)
            OTlist.append(OTdata)
        print(OTlist)
    elif themonth == july_input_sheet:
        month = july_input_sheet
        OTlist = []
        for d in range(1, 31):
            OTdata = OTperday(4 + d*2)
            OTlist.append(OTdata)
        print(OTlist)

print(OTlist)

"""
Basically with the code here I am trying to design it so that every month is generated automatically, 
and so that Haley does not have to manually enter in the month.
"""