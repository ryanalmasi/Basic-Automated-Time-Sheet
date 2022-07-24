"""
The purpose of this first part of the code is to combine the data from everyones timesheets. 
Note that a lot of the code currently on the sheet will just do a specific function, as printing a cell from 
an excel sheet. Will soon start to tie everything together
"""
# imports 
from multiprocessing.reduction import ForkingPickler
from pathlib import Path
from re import T
from types import NoneType
import webbrowser
from numpy import true_divide
import pandas as pd
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import glob

# to read a single cell from excel file with pandas
def read_value_from_excel(filename, column="C", row=3):
    return pd.read_excel(filename, skiprows=row - 1, usecols=column, nrows=1, header=None, names=["Value"]).iloc[0]["Value"]

# create a workbook object
# wb = Workbook()

# load existing spreadsheets
# note for when we are using files in a different directory we will have to do: ('C:/whateverthepathis/test.xlsx)
#wb = load_workbook('test.xlsx')
# wb1 = load_workbook('TS-2022_AbhroChowdhury.xlsx') ---------------------------------------

#create an active worksheet
#ws = wb.active
# ws1 = wb1.active-------------------------------

# to print something from our spreadsheet
#print(ws['A1'].value)
# print(ws1['Q4'].value) ----------------------------

# to print multiple things from spreadsheet we can use an f-string
#print(f'{ws["A2"].value} {ws["A1"].value:}')

# Set a variable to do it --> I can do this to pull the number of OT hours in the excel file and add them to a list
"""
mylist = []
name = ws["A3"].value
name2 = ws["C1"].value
mylist.append(name)
mylist.append(name2)
print(mylist)
print(sum(mylist))
print(f'{name}')
"""

# To specify which sheet to pull data from
# month_sheet = wb1['Jun'] ---------------------------------------
# nameofthesheet = (f'{month_sheet}')   --> trying to find a way to assign the sheet name to a variable; Then we can utilise the variable sheet name in the final excel sheet
# print(nameofthesheet)
# To loop through data
"""
is_data = True
count = 1
while is_data:
    count+=1
    data = month_sheet.cell(row=1, column=count).value
    if data == None:
        is_data = False
print(count)
"""

# to pull data from sheets, put it into a list, and how to add it all together
"""
datalist = []
for i in range(1, count):
    data = {}
    data = month_sheet.cell(row=1, column=i).value
    datalist.append(data)

print(datalist)
x = sum(datalist)
print("This months overtime is: ", x)
"""

""""
Potential solution to skipping every second cell. Create function containing for loop written above. 
Use it for every column containing OT entries, add to the same list, and sum it all up at the end. 
Or make list for each day, add it to that list, then we can input that list into wherever necessary
"""

"""
# to iterate over each day and count overtime hours
whilebreaker = True
count = 1
while whilebreaker:
    count += 1
    data1 = month_sheet.cell(row=count, column=16)
    if data1 == None:
        whilebreaker = True

abhrolist = []
for i in range(11, 70):
    data1 = {}
    data1 = month_sheet.cell(row=11+i, column=16).value
    abhrolist.append(data1)

print(abhrolist)
print(sum(abhrolist))
"""

# try this next
wb = load_workbook("TS-2022_AbhroChowdhury.xlsx")   # loading excel book
ws = wb.active   # loading excel sheet
enter_a_month = 'Dec'   # for now Haley will need to enter whatever month she needs it for
enter_OT_month = 'DEC 22'    # this has to follow {MONTH 22} format
month = wb[enter_a_month]     # determines what sheet to work on 
# print(ws['F12'].value)   --> to test cell location
# print(type(ws['F12'].value))    --> to test cell type (has to be int)

""" (All the following code is the body of the function)
mylist = []    # creating an empty list to later append values to 
for i in range(1, 20):    # had to put range 5 even though there are only 4 slots
    data = {}   # creating dictionary 
    data = month.cell(row=i, column=3).value     # gives me value for the column we want, and will iterate over each row in that column
    if data != None:   # if the cell is blank, we do not want to add it into the list 
        mylist.append(data)    # adding values into the list, this will give us the final overtime hours in that particular day

print(mylist)   
sum = sum(mylist) 
print("The sum of the list is: ", sum)
# This is a way to add all the OT hours in a particular column, while skipping the empty rows
# By doing this, I can get the final OT for each day. I will need to find a way to make the code do this for each OT column tho.
# I will create a function containing the above code, where one of the inputs is the column number
"""

def OTperday(specialcolumn):
    mylist = []    # creating an empty list to later append values to 
    for i in range(11, 20):    # range of rows that I want to iterate over
        checktype = type(month.cell(row=i, column=specialcolumn).value)
        if checktype == int:   # I want to only add integers to my list
            data = month.cell(row=i, column=specialcolumn).value     # gives me value for the column we want, and will iterate over each row in that column
            mylist.append(data)    # adding values into the list, this will give us the final overtime hours in that particular day
    thesum = sum(mylist)
    return thesum

OTlist = []
for d in range(1, 31):
    OTdata = OTperday(4 + d*2)
    OTlist.append(OTdata)


Dec01 = OTperday(6)   # Will tell me the overtime hours for that particular day
Dec02 = OTperday(8)
Dec05 = OTperday(14)


"""
Long way of making the OTlist (fixed now)
Dec01 = OTperday(6)   # Will tell me the overtime hours for that particular day
Dec02 = OTperday(8)
Dec03 = OTperday(10)
Dec04 = OTperday(12)
Dec05 = OTperday(14)
Dec06 = OTperday(16)
Dec07 = OTperday(18)
Dec08 = OTperday(20)
Dec09 = OTperday(22)
Dec10 = OTperday(24)
Dec11 = OTperday(26)
Dec12 = OTperday(28)
Dec13 = OTperday(30)
Dec14 = OTperday(32)
Dec15 = OTperday(34)
Dec16 = OTperday(36)
Dec17 = OTperday(38)
Dec18 = OTperday(40)
Dec19 = OTperday(42)
Dec20 = OTperday(44)
Dec21 = OTperday(46)
Dec22 = OTperday(48)
Dec23 = OTperday(50)
Dec24 = OTperday(52)
Dec25 = OTperday(54)
Dec26 = OTperday(56)
Dec27 = OTperday(58)
Dec28 = OTperday(60)
Dec29 = OTperday(62)
Dec30 = OTperday(64)
Dec31 = OTperday(66)
OTListDecember = [ Dec01, Dec02, Dec03, Dec04, Dec05, Dec06, Dec07, Dec08, Dec09, Dec10, Dec11, Dec12, Dec13, Dec14,
    Dec15, Dec16, Dec17, Dec18, Dec19, Dec20, Dec21, Dec22, Dec23, Dec24, Dec25, Dec26, Dec27, Dec28, Dec29, Dec30, Dec21,]
"""

# Figuring out how to write to an existing excel sheet; will first try to write to test.xlsx then the actual template
file_path = 'OT_template.xlsx'   # initiating the file we want to write to
wb = load_workbook(file_path)    # loading workbook with this file
ws = wb[enter_OT_month]   # uses input enter a month variable to determine the active sheet (so Haley doesn't have to retype it often)
ws['B12'] = Dec01
wb.save(file_path)



# figuring out how to automate writing the date
wb = load_workbook('test.xlsx')
ws = wb['Sheet1']
superlist = ["Dec01", "Dec02", "Dec03", "Dec04"]

def write_row(write_sheet, row_num: int, starting_column: int, write_values: list):
    for i, value in enumerate(write_values):
        write_sheet.cell(row_num + i, starting_column, value)

write_row(ws, 5, 1, superlist)
wb.save('test.xlsx')

for k in range(1,31):
    if enter_a_month == "Dec":
        month12 = "December"
        kstring1 = str(k)
        month12list = (f"December" + {kstring1} + ", 2022")
        print(month12list)


