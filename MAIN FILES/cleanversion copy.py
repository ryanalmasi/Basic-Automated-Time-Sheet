# imports 
import os
from pathlib import Path
from re import T
from tkinter import END, Y
from types import NoneType
import webbrowser
from numpy import full, true_divide
import pandas as pd
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import glob

"""
Initializations
"""
directory = r"C:\Users\ryana\Desktop\FolderForCode"
timesheet_month = ["Jan", "Feb", "Mar", "Apr", "May", "Jun", "Jul", "Aug", "Sep", "Oct", "Nov", "Dec"]  # Array of months for input file specification.
outputsheet_month = ["JAN 22", "FEB 22", "MAR 22", "APR 22", "MAY 22", "JUN 22", "JUL 22", "AUG 22", "SEP 22", "OCT 22", "NOV 22", "DEC 22"]    # Array of months for ouput file specification.
full_month = ["January", "February", "March", "April", "May", "June", "July", "August", "September", "October", "November", "December"]     # Array of full months to paste onto output template.



"""
Looping through Folder
"""
name = input("Name: ")
for filename in os.listdir(directory):
    if filename.endswith(".xlsx"):
        timesheet_wb = load_workbook(f"TS-2022_{name}.xlsx")   # loading excel book
        ws = timesheet_wb.active   # loading excel sheet
        output_file = f"OT_template{name}.xlsx"    # saving the output file name for later
        output_workbook = load_workbook(output_file)   # loading in the output file
        break

"""
This is the main part of the code; This function will pull how many overtime hours is worked per day and then store that data.
The function itself is complete and working, however based on how we can pull up the other excel files, we'll need to find
a way to then store that data for the third step
"""
def OTperday(specialcolumn):
    mylist = []    # creating an empty list to later append values to 
    for i in range(11, 30):    # range of rows that I want to iterate over (usually 11 to 70)
        checktype = type(month.cell(row=i, column=specialcolumn).value)  # this will pick out the specific cell, and check the type of value in it
        if checktype == int:   # I want to only add integers to my list
            data = month.cell(row=i, column=specialcolumn).value     # gives me value for the column we want, and will iterate over each row in that column
            mylist.append(data)    # adding values into the list, this will give us the final overtime hours in that particular day
    thesum = sum(mylist)    # taking total of the list to calculate total OT for that particular day
    return thesum


"""
This function will populate the dates for the sheet we are working on 
"""
def write_row(write_sheet, row_num: int, starting_column: int, write_values: list):    # row and column will always stay same, just change the list to match month
    for i, value in enumerate(write_values):    # loop over each row in the same column 
        write_sheet.cell(row_num + i, starting_column, value)     # writes to each row in the column


"""
To loop over initial input excel document in order to copy necessary data and paste onto template sheet.
"""
x = 0
while x < len(timesheet_month):
    output_ws = output_workbook[outputsheet_month[x]]
    month = timesheet_wb[timesheet_month[x]]
    OTlist = []
    for d in range(1, 31):
        OTdata = OTperday(4 + d*2)
        OTlist.append(OTdata)
        write_row(output_ws, 12, 2, OTlist)
    x += 1


"""
This third portion is to now write the data to an existing excel file. It will be easiest to write it to a template document
rather than creating a new document and formatting it from start
"""
y = 0
while y < len(full_month):
    output_ws = output_workbook[outputsheet_month[y]]
    Cal_list = []
    for day in range(1, 32):
        dayString = str(day)
        calendar = (str(full_month[y]) + " " + dayString + ", 2022")
        Cal_list.append(calendar)
        write_row(output_ws, 12, 1, Cal_list)   
    y += 1

output_workbook.save(output_file)
