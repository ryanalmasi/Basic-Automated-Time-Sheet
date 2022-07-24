# imports 
from pathlib import Path
from re import T
from types import NoneType
import webbrowser
from numpy import true_divide
import pandas as pd
from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import glob


"""
This is the initial loading phase using openpyxl, currently works for 1 file however we need a way to load many files
"""
timesheet_wb = load_workbook("TS-2022_AbhroChowdhury.xlsx")   # loading excel book
ws = timesheet_wb.active   # loading excel sheet
timesheet_month = "Jan"    # Haley will enter the timesheet month here
outputsheet_month = "JAN 22"    # Haley must enter this too, format is: {MONTH 22}
month = timesheet_wb[timesheet_month]     # determines what sheet to work on 
output_file = 'OT_template.xlsx'    # saving the output file name for later
output_workbook = load_workbook(output_file)   # loading in the output file
output_ws = output_workbook[outputsheet_month]    # loading in the month in the output file
output_ws = output_workbook[outputsheet_month]    # loading in the month in the output file


"""
This is the main part of the code; This function will pull how many overtime hours is worked per day and then store that data.
The function itself is complete and working, however based on how we can pull up the other excel files, we'll need to find
a way to then store that data for the third step
"""
def OTperday(specialcolumn):
    mylist = []    # creating an empty list to later append values to 
    for i in range(11, 30):    # range of rows that I want to iterate over (usually 11 to 70)
        checktype = type(month.cell(row=i, column=specialcolumn).value)  # this will pick out the specific cell, and check the type of value in it
        if checktype == int:   # I want to only add integers to my list
            data = month.cell(row=i, column=specialcolumn).value     # gives me value for the column we want, and will iterate over each row in that column
            mylist.append(data)    # adding values into the list, this will give us the final overtime hours in that particular day
    thesum = sum(mylist)    # taking total of the list to calculate total OT for that particular day
    return thesum

OTlist = []
for d in range(1, 31):
    OTdata = OTperday(4 + d*2)
    OTlist.append(OTdata)


"""
This third portion is to now write the data to an existing excel file. It will be easiest to write it to a template document
rather than creating a new document and formatting it from start
"""
# This function will populate the dates for the sheet we are working on 
def write_row(write_sheet, row_num: int, starting_column: int, write_values: list):    # row and column will always stay same, just change the list to match month
    for i, value in enumerate(write_values):    # loop over each row in the same column 
        write_sheet.cell(row_num + i, starting_column, value)     # writes to each row in the column

# Takes month entered by Haley, creates dates for that month, appends it to a list, then prints that list to the corresponding rows
if timesheet_month == "Jan":
    JanuaryList = []
    for day in range(1,32):
        if timesheet_month == "Jan":
            month1 = "January"
            daystring1 = str(day)
            jan_calendar = (month1 + ' ' + daystring1 + ", 2022")
            JanuaryList.append(jan_calendar)
        write_row(output_ws, 12, 1, JanuaryList)   
        write_row(output_ws, 12, 2, OTlist)
        output_workbook.save(output_file)
elif timesheet_month == "Feb":
    FebruaryList = []
    for day in range(1,32):
        if timesheet_month == "Feb":
            month2 = "February"
            daystring2 = str(day)
            feb_calendar = (month2 + ' ' + daystring2 + ", 2022")
            FebruaryList.append(feb_calendar)
        write_row(output_ws, 12, 1, FebruaryList) 
        write_row(output_ws, 12, 2, OTlist)
        output_workbook.save(output_file)
elif timesheet_month == "Mar":
    MarchList = []
    for day in range(1,32):
        if timesheet_month == "Mar":
            month3 = "March"
            daystring3 = str(day)
            mar_calendar = (month3 + ' ' + daystring3 + ", 2022")
            MarchList.append(mar_calendar)
        write_row(output_ws, 12, 1, MarchList)   
        write_row(output_ws, 12, 2, OTlist)
        output_workbook.save(output_file)
elif timesheet_month == "Apr":
    AprilList = []
    for day in range(1,32):
        if timesheet_month == "Apr":
            month4 = "April"
            daystring4 = str(day)
            apr_calendar = (month4 + ' ' + daystring4 + ", 2022")
            AprilList.append(apr_calendar)
        write_row(output_ws, 12, 1, AprilList)  
        write_row(output_ws, 12, 2, OTlist)
        output_workbook.save(output_file)
elif timesheet_month == "May":
    MayList = []
    for day in range(1,32):
        if timesheet_month == "May":
            month5 = "May"
            daystring5 = str(day)
            may_calendar = (month5 + ' ' + daystring5 + ", 2022")
            MayList.append(may_calendar)
        write_row(output_ws, 12, 1, MayList)  
        write_row(output_ws, 12, 2, OTlist)
        output_workbook.save(output_file)
elif timesheet_month == "Jun":
    JuneList = []
    for day in range(1,32):
        if timesheet_month == "Jun":
            month6 = "Jun"
            daystring6 = str(day)
            jun_calendar = (month6 + '' + daystring6 + ", 2022")
            JuneList.append(jun_calendar)
        write_row(output_ws, 12, 1, JuneList)  
        write_row(output_ws, 12, 2, OTlist)
        output_workbook.save(output_file)
elif timesheet_month == "July":
    JulyList = []
    for day in range(1,32):
        if timesheet_month == "July":
            month7 = "July"
            daystring7 = str(day)
            jul_calendar = (month7 + ' ' + daystring7 + ", 2022")
            JulyList.append(jul_calendar)
        write_row(output_ws, 12, 1, JulyList)   
        write_row(output_ws, 12, 2, OTlist)
        output_workbook.save(output_file)
elif timesheet_month == "Aug":
    AugustList = []
    for day in range(1,32):
        if timesheet_month == "Aug":
            month8 = "August"
            daystring8 = str(day)
            aug_calendar = (month8 + ' ' + daystring8 + ", 2022")
            AugustList.append(aug_calendar)
        write_row(output_ws, 12, 1, AugustList)    
        write_row(output_ws, 12, 2, OTlist)
        output_workbook.save(output_file)
elif timesheet_month == "Sep":
    SeptemberList = []
    for day in range(1,32):
        if timesheet_month == "Sep":
            month9 = "September"
            daystring9 = str(day)
            sep_calendar = (month9 + ' ' + daystring9 + ", 2022")
            SeptemberList.append(sep_calendar)
        write_row(output_ws, 12, 1, SeptemberList)   
        write_row(output_ws, 12, 2, OTlist)
        output_workbook.save(output_file)
elif timesheet_month == "Oct":
    OctoberList = []
    for day in range(1,32):
        if timesheet_month == "Oct":
            month10 = "October"
            daystring10 = str(day)
            oct_calendar = (month10 + ' ' + daystring10 + ", 2022")
            OctoberList.append(oct_calendar)
        write_row(output_ws, 12, 1, OctoberList)  
        write_row(output_ws, 12, 2, OTlist)
        output_workbook.save(output_file)
elif timesheet_month == "Nov":
    NovemberList = []
    for day in range(1,32):
        if timesheet_month == "Nov":
            month11 = "November"
            daystring11 = str(day)
            nov_calendar = (month11 + ' ' + daystring11 + ", 2022")
            NovemberList.append(nov_calendar)
        write_row(output_ws, 12, 1, NovemberList)   
        write_row(output_ws, 12, 2, OTlist)
        output_workbook.save(output_file)
elif timesheet_month == "Dec":
    DecemberList = []
    for day in range(1,32):
        if timesheet_month == "Dec":
            month12 = "December"
            daystring12 = str(day)
            dec_calendar = (month12 + ' ' + daystring12 + ", 2022")
            DecemberList.append(dec_calendar)
        write_row(output_ws, 12, 1, DecemberList)  
        write_row(output_ws, 12, 2, OTlist)
        output_workbook.save(output_file)







